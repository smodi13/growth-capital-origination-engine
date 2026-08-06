#!/usr/bin/env python3
"""
Build Enterprise_Software_Growth_Capital_Model.xlsx.

Design rules for this workbook:

  * Every output is a live formula referencing the Assumptions sheet. Change an
    assumption and the whole workbook, including the returns and the
    recommendation logic, recalculates. Nothing is a pasted number.
  * No macros. openpyxl only, .xlsx format.
  * Conditional formatting is used sparingly and only where it carries meaning:
    covenant breaches and negative EBITDA.
  * Every output sheet carries the line "All figures are hypothetical and
    illustrative."

Run: python3 scripts/build_model.py
"""

from __future__ import annotations

import os
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "public",
    "downloads",
    "Enterprise_Software_Growth_Capital_Model.xlsx",
)

COMPANY = "Northstar Workflow Systems"
DISCLOSURE = "All figures are hypothetical and illustrative."
LONG_DISCLOSURE = (
    f"{COMPANY} is a hypothetical company created solely for an illustrative "
    "underwriting exercise. It is not a real business or investment opportunity."
)

# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------

INK = "0F141C"
HEAD_FILL = PatternFill("solid", fgColor="1A2230")
BAND_FILL = PatternFill("solid", fgColor="F2F5F8")
INPUT_FILL = PatternFill("solid", fgColor="FFF7E0")
TOTAL_FILL = PatternFill("solid", fgColor="E8EEF3")
ACCENT = "0F5F66"

THIN = Side(style="thin", color="C9D2DC")
MED = Side(style="medium", color="8895A5")

FMT_M = '#,##0.0;(#,##0.0)'
FMT_M2 = '#,##0.00;(#,##0.00)'
FMT_PCT = '0.0%;(0.0%)'
FMT_PCT0 = '0%'
FMT_X = '0.00"x"'
FMT_INT = '#,##0'
FMT_K = '#,##0'


def styles(wb: Workbook) -> None:
    title = NamedStyle(name="title")
    title.font = Font(name="Calibri", size=16, bold=True, color=INK)

    sub = NamedStyle(name="sub")
    sub.font = Font(name="Calibri", size=10, italic=True, color="5B6878")

    hdr = NamedStyle(name="hdr")
    hdr.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    hdr.fill = HEAD_FILL
    hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hdr.border = Border(bottom=MED)

    lbl = NamedStyle(name="lbl")
    lbl.font = Font(name="Calibri", size=10, color=INK)
    lbl.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    sec = NamedStyle(name="sec")
    sec.font = Font(name="Calibri", size=11, bold=True, color=ACCENT)
    sec.border = Border(bottom=THIN)

    inp = NamedStyle(name="inp")
    inp.font = Font(name="Calibri", size=10, bold=True, color="7A4E00")
    inp.fill = INPUT_FILL
    inp.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    inp.alignment = Alignment(horizontal="right")

    calc = NamedStyle(name="calc")
    calc.font = Font(name="Calibri", size=10, color=INK)
    calc.alignment = Alignment(horizontal="right")

    tot = NamedStyle(name="tot")
    tot.font = Font(name="Calibri", size=10, bold=True, color=INK)
    tot.fill = TOTAL_FILL
    tot.border = Border(top=THIN, bottom=MED)
    tot.alignment = Alignment(horizontal="right")

    note = NamedStyle(name="note")
    note.font = Font(name="Calibri", size=9, italic=True, color="5B6878")
    note.alignment = Alignment(wrap_text=True, vertical="top")

    body = NamedStyle(name="body")
    body.font = Font(name="Calibri", size=10, color=INK)
    body.alignment = Alignment(wrap_text=True, vertical="top")

    for s in (title, sub, hdr, lbl, sec, inp, calc, tot, note, body):
        wb.add_named_style(s)


def sheet_header(ws, title: str, subtitle: str, last_col: str = "H") -> int:
    """Write the standard title block. Returns the next free row."""
    ws["A1"] = title
    ws["A1"].style = "title"
    ws["A2"] = subtitle
    ws["A2"].style = "sub"
    ws["A3"] = DISCLOSURE
    ws["A3"].style = "sub"
    ws["A3"].font = Font(name="Calibri", size=10, italic=True, bold=True, color="9C4221")
    ws.merge_cells(f"A1:{last_col}1")
    ws.merge_cells(f"A2:{last_col}2")
    ws.merge_cells(f"A3:{last_col}3")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "B6"
    return 5


def year_headers(ws, row: int, first_col: int = 2, n: int = 5, label: str = "Line item"):
    ws.cell(row=row, column=1, value=label).style = "hdr"
    for i in range(n):
        c = ws.cell(row=row, column=first_col + i, value=f"Year {i + 1}")
        c.style = "hdr"
    return row + 1


def widths(ws, first: int, rest: int, n: int = 8):
    ws.column_dimensions["A"].width = first
    for i in range(2, 2 + n):
        ws.column_dimensions[get_column_letter(i)].width = rest


# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------

wb = Workbook()
styles(wb)

# =============================================================== 1. Read Me ==
ws = wb.active
ws.title = "Read Me"
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 108

ws["A1"] = "Enterprise Software Growth Capital Model"
ws["A1"].style = "title"
ws["A2"] = f"Illustrative underwriting of {COMPANY}"
ws["A2"].style = "sub"

r = 4
readme = [
    ("Disclosure", LONG_DISCLOSURE),
    ("Purpose",
     "This workbook underwrites a single hypothetical B2B enterprise SaaS company across three capital "
     "structures: all growth equity, all private credit, and a blended structure combining both. The "
     "operating forecast is identical in every case, so any difference in outcome is caused by the "
     "structure rather than by the business."),
    ("How to use it",
     "Every input lives on the Assumptions sheet and is shaded amber. Change any amber cell and the entire "
     "workbook recalculates, including the SaaS metrics, the debt schedules, the returns, the sensitivities, "
     "and the downside case. No output anywhere in this workbook is a hardcoded number."),
    ("Colour convention",
     "Amber cells are inputs you may change. White cells are formulas driven by those inputs. Grey banded "
     "rows are subtotals. Red shading appears only where a minimum cash covenant is breached or EBITDA is "
     "negative, so the formatting carries meaning rather than decoration."),
    ("Sheet order",
     "Read Me, Assumptions, Historical Financials, Operating Forecast, SaaS Metrics, Capital Structures, "
     "Growth Equity Case, Private Credit Case, Blended Capital Case, Debt Schedule, Returns Analysis, "
     "Sensitivities, Downside Case, Sources and Disclosures."),
    ("Key formula: ARR bridge",
     "Ending ARR = Beginning ARR + New ARR + Expansion ARR less Churned ARR. Churned ARR is set by gross "
     "retention and Expansion ARR by the gap between net and gross retention. New ARR is then the residual "
     "required to hit the growth target, which makes the sales requirement explicit rather than assumed. "
     "If you lower net revenue retention, required New ARR rises automatically."),
    ("Key formula: Revenue",
     "Revenue = average of beginning and ending ARR for the year. ARR is added through the year rather than "
     "on day one, so recognising the full ending ARR as revenue would overstate the period."),
    ("Key formula: Unlevered free cash flow",
     "EBITDA less capital expenditure plus the deferred revenue benefit on ARR growth. It is stated before "
     "any interest or principal so the same cash flow line can be used across all three structures."),
    ("Key formula: Debt service",
     "Cash interest on the opening balance, plus amortisation once the interest only period ends. PIK "
     "interest, if switched on, accretes to the balance rather than being paid in cash."),
    ("Key formula: Exit equity value",
     "Exit enterprise value at the exit ARR multiple, plus ending cash, less ending debt. The equity "
     "investor receives their ownership percentage of that."),
    ("What this workbook is not",
     "It is not a valuation of any real company, not an investment recommendation, and not a forecast. It "
     "is a demonstration of how a structure decision changes dilution, debt service risk, covenant "
     "headroom, runway, and returns when the underlying business is held constant."),
]
for label, text in readme:
    ws.cell(row=r, column=1, value=label).style = "sec"
    c = ws.cell(row=r, column=2, value=text)
    c.style = "body"
    ws.row_dimensions[r].height = max(30, 13 * (len(text) // 110 + 1))
    r += 1

ws.cell(row=r + 1, column=1, value="Built by").style = "sec"
ws.cell(row=r + 1, column=2, value="Sahil Modi, modi.sahil@gmail.com").style = "body"

# ============================================================ 2. Assumptions ==
aw = wb.create_sheet("Assumptions")
widths(aw, 46, 15, 8)
r = sheet_header(aw, "Assumptions", "Every amber cell is an input. Everything else in the workbook derives from these.", "G")


def block(ws, row, title):
    ws.cell(row=row, column=1, value=title).style = "sec"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    return row + 1


def inp(ws, row, label, value, fmt, comment=None, unit=""):
    ws.cell(row=row, column=1, value=label).style = "lbl"
    c = ws.cell(row=row, column=2, value=value)
    c.style = "inp"
    c.number_format = fmt
    if unit:
        u = ws.cell(row=row, column=3, value=unit)
        u.style = "note"
    if comment:
        c.comment = Comment(comment, "Model guide")
    return row + 1


r = block(aw, r, "Operating assumptions")
A = {}
r = inp(aw, r, "Beginning ARR (USD millions)", 12.0, FMT_M2,
        "Starting annual recurring revenue. Drives the entire ARR bridge on the Operating Forecast sheet.")
A["beg_arr"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Year 1 ARR growth rate", 0.30, FMT_PCT,
        "Year one growth. Later years step down by the growth decay below.")
A["g1"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Annual growth decay (percentage points)", 0.02, FMT_PCT,
        "Subtracted from the growth rate each year. Set to 0 to hold growth flat.")
A["decay"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Net revenue retention", 1.10, FMT_PCT,
        "Expansion less churn within the existing base. Raising this lowers the New ARR the sales team must win.")
A["nrr"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Gross retention", 0.88, FMT_PCT,
        "Share of beginning ARR retained before expansion. 1 less this figure is the churn rate.")
A["grr"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Gross margin", 0.78, FMT_PCT,
        "Applied to revenue to give gross profit.")
A["gm"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Capital expenditure (% of revenue)", 0.02, FMT_PCT)
A["capex"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Deferred revenue benefit (% of ARR increase)", 0.10, FMT_PCT,
        "Working capital source from customers paying annually in advance as ARR grows.")
A["dr"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Beginning cash (USD millions)", 8.0, FMT_M2)
A["cash0"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Existing debt (USD millions)", 3.0, FMT_M2,
        "Refinanced in full at close in every structure.")
A["debt0"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Customer count at year 0", 240, FMT_INT)
A["cust0"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "New customer ACV premium to base", 1.15, FMT_X,
        "New customers are won at a higher average contract value than the existing base average.")
A["acvprem"] = f"Assumptions!$B${r-1}"

r += 1
r = block(aw, r, "Operating expense ratios (% of revenue)")
opex_rows = {}
aw.cell(row=r, column=1, value="Line").style = "hdr"
for i in range(5):
    aw.cell(row=r, column=2 + i, value=f"Year {i+1}").style = "hdr"
r += 1
for name, vals in (
    ("Sales and marketing", [0.55, 0.52, 0.48, 0.44, 0.40]),
    ("Research and development", [0.30, 0.28, 0.26, 0.24, 0.22]),
    ("General and administrative", [0.18, 0.16, 0.15, 0.14, 0.13]),
):
    aw.cell(row=r, column=1, value=name).style = "lbl"
    for i, v in enumerate(vals):
        c = aw.cell(row=r, column=2 + i, value=v)
        c.style = "inp"
        c.number_format = FMT_PCT
    opex_rows[name] = r
    r += 1
A["sm_row"], A["rd_row"], A["ga_row"] = (
    opex_rows["Sales and marketing"],
    opex_rows["Research and development"],
    opex_rows["General and administrative"],
)

r += 1
r = block(aw, r, "Transaction assumptions")
r = inp(aw, r, "Total capital raised (USD millions)", 20.0, FMT_M2)
A["raise"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Existing debt refinanced at close (USD millions)", 3.0, FMT_M2)
A["refi"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Entry ARR multiple (pre-money)", 8.0, FMT_X,
        "Pre-money valuation equals this multiple times beginning ARR.")
A["entryx"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Exit ARR multiple", 7.0, FMT_X,
        "Applied to year five ending ARR. Sensitised on the Sensitivities sheet.")
A["exitx"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Minimum cash covenant (USD millions)", 5.0, FMT_M2,
        "Cash balances below this level are flagged red on the case sheets.")
A["mincash"] = f"Assumptions!$B${r-1}"

r += 1
r = block(aw, r, "Growth equity case")
r = inp(aw, r, "New equity invested (USD millions)", 20.0, FMT_M2)
A["eq_equity"] = f"Assumptions!$B${r-1}"

r += 1
r = block(aw, r, "Private credit case")
r = inp(aw, r, "Principal (USD millions)", 20.0, FMT_M2)
A["pc_principal"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Cash interest rate", 0.115, FMT_PCT)
A["pc_cash"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "PIK interest rate", 0.0, FMT_PCT,
        "Optional. PIK accretes to the balance rather than being paid in cash. Set above zero to test it.")
A["pc_pik"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Original issue discount", 0.02, FMT_PCT,
        "Reduces net proceeds at close and raises the lender return.")
A["pc_oid"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Maturity (years)", 5, FMT_INT)
A["pc_mat"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Interest only period (years)", 3, FMT_INT)
A["pc_io"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Annual amortisation (% of original principal)", 0.05, FMT_PCT)
A["pc_amort"] = f"Assumptions!$B${r-1}"

r += 1
r = block(aw, r, "Blended capital case")
r = inp(aw, r, "Equity component (USD millions)", 8.0, FMT_M2)
A["bl_equity"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Debt component (USD millions)", 12.0, FMT_M2)
A["bl_debt"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Cash interest rate", 0.11, FMT_PCT)
A["bl_cash"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "PIK interest rate", 0.0, FMT_PCT)
A["bl_pik"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Original issue discount", 0.015, FMT_PCT)
A["bl_oid"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Maturity (years)", 5, FMT_INT)
A["bl_mat"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Interest only period (years)", 3, FMT_INT)
A["bl_io"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Annual amortisation (% of original principal)", 0.05, FMT_PCT)
A["bl_amort"] = f"Assumptions!$B${r-1}"

r += 1
r = block(aw, r, "Downside case overrides")
r = inp(aw, r, "Year 1 ARR growth rate", 0.18, FMT_PCT)
A["ds_g1"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Net revenue retention", 1.00, FMT_PCT)
A["ds_nrr"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Gross retention", 0.82, FMT_PCT)
A["ds_grr"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Gross margin", 0.74, FMT_PCT)
A["ds_gm"] = f"Assumptions!$B${r-1}"
r = inp(aw, r, "Sales and marketing uplift (percentage points)", 0.05, FMT_PCT,
        "Sales and marketing does not fall as quickly when growth disappoints.")
A["ds_smup"] = f"Assumptions!$B${r-1}"

aw.cell(row=r + 1, column=1, value=DISCLOSURE).style = "note"

# ==================================================== 3. Historical Financials ==
hw = wb.create_sheet("Historical Financials")
widths(hw, 40, 15, 4)
r = sheet_header(hw, "Historical Financials", f"Illustrative three year history for {COMPANY}, ending at the model start point.", "E")

hw.cell(row=r, column=1, value="Line item (USD millions)").style = "hdr"
for i, y in enumerate(["Year -3", "Year -2", "Year -1", "Year 0"]):
    hw.cell(row=r, column=2 + i, value=y).style = "hdr"
r += 1

hist_start = r
hist = [
    ("Ending ARR", [4.9, 6.9, 9.4, None], FMT_M2, "arr"),
    ("ARR growth", [None, None, None, None], FMT_PCT, "growth"),
    ("Revenue", [None, None, None, None], FMT_M2, "rev"),
    ("Gross profit", [None, None, None, None], FMT_M2, "gp"),
    ("Total operating expenses", [5.4, 7.3, 9.6, 12.7], FMT_M2, "opex"),
    ("EBITDA", [None, None, None, None], FMT_M2, "ebitda"),
    ("Cash balance", [6.2, 9.8, 7.1, None], FMT_M2, "cash"),
    ("Customer count", [118, 156, 199, None], FMT_INT, "cust"),
]
rows = {}
for name, vals, fmt, key in hist:
    hw.cell(row=r, column=1, value=name).style = "lbl"
    rows[key] = r
    r += 1

# ARR row: last column links to the Assumptions beginning ARR.
for i, v in enumerate([4.9, 6.9, 9.4]):
    c = hw.cell(row=rows["arr"], column=2 + i, value=v)
    c.style = "inp"
    c.number_format = FMT_M2
c = hw.cell(row=rows["arr"], column=5, value=f"={A['beg_arr']}")
c.style = "calc"
c.number_format = FMT_M2

for i in range(1, 4):
    col = get_column_letter(2 + i)
    prev = get_column_letter(1 + i)
    c = hw.cell(row=rows["growth"], column=2 + i,
                value=f"={col}{rows['arr']}/{prev}{rows['arr']}-1")
    c.style = "calc"
    c.number_format = FMT_PCT
hw.cell(row=rows["growth"], column=2, value="n/a").style = "calc"

for i in range(4):
    col = get_column_letter(2 + i)
    if i == 0:
        c = hw.cell(row=rows["rev"], column=2, value=f"=B{rows['arr']}*0.85")
    else:
        prev = get_column_letter(1 + i)
        c = hw.cell(row=rows["rev"], column=2 + i,
                    value=f"=({prev}{rows['arr']}+{col}{rows['arr']})/2")
    c.style = "calc"
    c.number_format = FMT_M2

    c = hw.cell(row=rows["gp"], column=2 + i, value=f"={col}{rows['rev']}*{A['gm']}")
    c.style = "calc"
    c.number_format = FMT_M2

    c = hw.cell(row=rows["ebitda"], column=2 + i,
                value=f"={col}{rows['gp']}-{col}{rows['opex']}")
    c.style = "tot"
    c.number_format = FMT_M2

for i, v in enumerate([5.4, 7.3, 9.6, 12.7]):
    c = hw.cell(row=rows["opex"], column=2 + i, value=v)
    c.style = "inp"
    c.number_format = FMT_M2

for i, v in enumerate([6.2, 9.8, 7.1]):
    c = hw.cell(row=rows["cash"], column=2 + i, value=v)
    c.style = "inp"
    c.number_format = FMT_M2
c = hw.cell(row=rows["cash"], column=5, value=f"={A['cash0']}")
c.style = "calc"
c.number_format = FMT_M2

for i, v in enumerate([118, 156, 199]):
    c = hw.cell(row=rows["cust"], column=2 + i, value=v)
    c.style = "inp"
    c.number_format = FMT_INT
c = hw.cell(row=rows["cust"], column=5, value=f"={A['cust0']}")
c.style = "calc"
c.number_format = FMT_INT

hw.conditional_formatting.add(
    f"B{rows['ebitda']}:E{rows['ebitda']}",
    CellIsRule(operator="lessThan", formula=["0"], font=Font(color="9B1C1C")),
)

r += 1
hw.cell(row=r, column=1,
        value="History is illustrative and is provided so the forecast has a visible starting trajectory. "
              "Year 0 ARR, cash, and customer count link to the Assumptions sheet, so the history and the "
              "forecast always join.").style = "note"
hw.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
hw.row_dimensions[r].height = 40
hw.cell(row=r + 2, column=1, value=DISCLOSURE).style = "note"

# ====================================================== 4. Operating Forecast ==
ow = wb.create_sheet("Operating Forecast")
widths(ow, 44, 15, 6)
r = sheet_header(ow, "Operating Forecast",
                 "Five year base case. Churn and expansion are set by retention; New ARR is the residual required to hit the growth target.",
                 "F")
r = year_headers(ow, r, label="Line item (USD millions)")

O = {}


def orow(name, key, fmt=FMT_M2, style="calc", formula=None, band=False):
    """Add a forecast row. `formula` receives the 1-based year index and column letter."""
    global r
    ow.cell(row=r, column=1, value=name).style = "lbl"
    O[key] = r
    if formula:
        for i in range(5):
            col = get_column_letter(2 + i)
            c = ow.cell(row=r, column=2 + i, value=formula(i + 1, col))
            c.style = style
            c.number_format = fmt
    if band:
        for i in range(6):
            ow.cell(row=r, column=1 + i).fill = TOTAL_FILL
    r += 1


ow.cell(row=r, column=1, value="ARR bridge").style = "sec"
r += 1

# Beginning ARR needs to reference Ending ARR, which does not exist yet, so the
# row is reserved here and its formulas are written once Ending ARR is placed.
orow("Beginning ARR", "beg", formula=lambda i, c: f"={A['beg_arr']}")
_beg_row = O["beg"]

orow("Growth rate", "g", fmt=FMT_PCT, formula=lambda i, c: f"=MAX(0,{A['g1']}-{A['decay']}*{i-1})")
orow("New ARR", "new", formula=lambda i, c: f"={c}{O['g']}*0+0")   # placeholder, rewritten below
orow("Expansion ARR", "exp", formula=lambda i, c: f"={c}{_beg_row}*({A['nrr']}-{A['grr']})")
orow("Churned ARR", "chn", formula=lambda i, c: f"=-{c}{_beg_row}*(1-{A['grr']})")
orow("Ending ARR", "end", formula=lambda i, c: f"={c}{_beg_row}+{c}{O['new']}+{c}{O['exp']}+{c}{O['chn']}",
     style="tot", band=True)

# Rewrite the two rows that needed forward references.
for i in range(5):
    col = get_column_letter(2 + i)
    prev = get_column_letter(1 + i)
    c = ow.cell(row=_beg_row, column=2 + i,
                value=f"={A['beg_arr']}" if i == 0 else f"={prev}{O['end']}")
    c.style = "calc"
    c.number_format = FMT_M2
    # New ARR is the residual: target ending ARR less what the existing base retains.
    c = ow.cell(row=O["new"], column=2 + i,
                value=f"={col}{_beg_row}*(1+{col}{O['g']})-{col}{_beg_row}*{A['nrr']}")
    c.style = "calc"
    c.number_format = FMT_M2

ow.cell(row=O["new"], column=2).comment = Comment(
    "New ARR is a residual, not an input. It equals the ending ARR the growth target implies, less the "
    "ARR the existing base retains after churn and expansion. Lowering net revenue retention raises the "
    "New ARR the sales team must win to hit the same target.", "Model guide")

r += 1
ow.cell(row=r, column=1, value="Income statement").style = "sec"
r += 1
orow("Revenue", "rev", formula=lambda i, c: f"=({c}{_beg_row}+{c}{O['end']})/2")
orow("Gross profit", "gp", formula=lambda i, c: f"={c}{O['rev']}*{A['gm']}")
orow("Gross margin", "gmpct", fmt=FMT_PCT, formula=lambda i, c: f"={c}{O['gp']}/{c}{O['rev']}")
orow("Sales and marketing", "sm",
     formula=lambda i, c: f"=-{c}{O['rev']}*Assumptions!{c}${A['sm_row']}")
orow("Research and development", "rd",
     formula=lambda i, c: f"=-{c}{O['rev']}*Assumptions!{c}${A['rd_row']}")
orow("General and administrative", "ga",
     formula=lambda i, c: f"=-{c}{O['rev']}*Assumptions!{c}${A['ga_row']}")
orow("Total operating expenses", "opex",
     formula=lambda i, c: f"=SUM({c}{O['sm']}:{c}{O['ga']})")
orow("EBITDA", "ebitda", formula=lambda i, c: f"={c}{O['gp']}+{c}{O['opex']}", style="tot", band=True)
orow("EBITDA margin", "ebmargin", fmt=FMT_PCT, formula=lambda i, c: f"={c}{O['ebitda']}/{c}{O['rev']}")

r += 1
ow.cell(row=r, column=1, value="Cash flow").style = "sec"
r += 1
orow("Capital expenditure", "capex", formula=lambda i, c: f"=-{c}{O['rev']}*{A['capex']}")
orow("Deferred revenue benefit", "drb",
     formula=lambda i, c: f"=({c}{O['end']}-{c}{_beg_row})*{A['dr']}")
orow("Unlevered free cash flow", "fcf",
     formula=lambda i, c: f"={c}{O['ebitda']}+{c}{O['capex']}+{c}{O['drb']}", style="tot", band=True)

r += 1
ow.cell(row=r, column=1, value="Customers").style = "sec"
r += 1
# Same forward reference pattern as Beginning ARR: reserved here, filled below.
orow("Beginning customers", "bcust", fmt=FMT_INT, formula=lambda i, c: f"={A['cust0']}")
orow("New customers", "ncust", fmt=FMT_INT,
     formula=lambda i, c: f"={c}{O['new']}/(({c}{_beg_row}/{c}{O['bcust']})*{A['acvprem']})")
orow("Churned customers", "ccust", fmt=FMT_INT,
     formula=lambda i, c: f"=-{c}{O['bcust']}*(1-{A['grr']})")
orow("Ending customers", "ecust", fmt=FMT_INT,
     formula=lambda i, c: f"={c}{O['bcust']}+{c}{O['ncust']}+{c}{O['ccust']}", style="tot", band=True)

for i in range(5):
    col = get_column_letter(2 + i)
    prev = get_column_letter(1 + i)
    c = ow.cell(row=O["bcust"], column=2 + i,
                value=f"={A['cust0']}" if i == 0 else f"={prev}{O['ecust']}")
    c.style = "calc"
    c.number_format = FMT_INT

ow.conditional_formatting.add(
    f"B{O['ebitda']}:F{O['ebitda']}",
    CellIsRule(operator="lessThan", formula=["0"], font=Font(color="9B1C1C", bold=True)),
)

ow.cell(row=r + 1, column=1, value=DISCLOSURE).style = "note"

# ========================================================== 5. SaaS Metrics ==
sw = wb.create_sheet("SaaS Metrics")
widths(sw, 44, 15, 6)
r = sheet_header(sw, "SaaS Metrics",
                 "The metrics a growth capital or private credit investor tests first. All derived from the Operating Forecast.",
                 "F")
r = year_headers(sw, r, label="Metric")

S = {}


def srow(name, key, fmt, formula, comment=None, band=False):
    global r
    sw.cell(row=r, column=1, value=name).style = "lbl"
    S[key] = r
    for i in range(5):
        col = get_column_letter(2 + i)
        c = sw.cell(row=r, column=2 + i, value=formula(i + 1, col))
        c.style = "tot" if band else "calc"
        c.number_format = fmt
    if comment:
        sw.cell(row=r, column=2).comment = Comment(comment, "Model guide")
    r += 1


srow("ARR growth", "g", FMT_PCT, lambda i, c: f"='Operating Forecast'!{c}{O['g']}")
srow("Net revenue retention", "nrr", FMT_PCT, lambda i, c: f"={A['nrr']}")
srow("Gross retention", "grr", FMT_PCT, lambda i, c: f"={A['grr']}")
srow("Gross margin", "gm", FMT_PCT, lambda i, c: f"='Operating Forecast'!{c}{O['gmpct']}")
srow("CAC payback (months)", "cac", '0.0',
     lambda i, c: f"=-'Operating Forecast'!{c}{O['sm']}/(('Operating Forecast'!{c}{O['new']}+'Operating Forecast'!{c}{O['exp']})*{A['gm']})*12",
     "Blended CAC payback. Sales and marketing supports both new logos and expansion, so it is recovered "
     "against gross profit on new plus expansion ARR. Charging the whole line to new ARR alone would "
     "overstate payback materially.")
srow("Burn multiple", "burn", FMT_X,
     lambda i, c: f"=IF('Operating Forecast'!{c}{O['fcf']}>=0,\"n/a\",-'Operating Forecast'!{c}{O['fcf']}/('Operating Forecast'!{c}{O['end']}-'Operating Forecast'!{c}{O['beg']}))",
     "Net cash burned per dollar of net new ARR added. Below 1.0x is efficient; the ratio improves every "
     "year in the base case as operating leverage arrives.")
srow("Rule of 40", "r40", '0',
     lambda i, c: f"='Operating Forecast'!{c}{O['g']}*100+'Operating Forecast'!{c}{O['ebmargin']}*100",
     "ARR growth percentage plus EBITDA margin percentage. This case starts well below 40, which is the "
     "central weakness of the credit story and is stated as such in the memorandum.",
     band=True)
srow("Customer count", "cust", FMT_INT, lambda i, c: f"='Operating Forecast'!{c}{O['ecust']}")
srow("Average contract value (USD thousands)", "acv", FMT_K,
     lambda i, c: f"='Operating Forecast'!{c}{O['end']}/'Operating Forecast'!{c}{O['ecust']}*1000")
srow("Revenue per customer (USD thousands)", "rpc", FMT_K,
     lambda i, c: f"='Operating Forecast'!{c}{O['rev']}/'Operating Forecast'!{c}{O['ecust']}*1000")

sw.conditional_formatting.add(
    f"B{S['r40']}:F{S['r40']}",
    CellIsRule(operator="lessThan", formula=["40"], font=Font(color="9B1C1C")),
)

r += 1
sw.cell(row=r, column=1,
        value="Where this case is strong: net revenue retention above 100 percent means the installed base grows "
              "without new logos, and the burn multiple improves every year. Where it is weak: Rule of 40 never "
              "reaches 40 inside the forecast, and gross retention of 88 percent means 12 percent of the base is "
              "lost annually. Gross retention is the single assumption most worth challenging in diligence.").style = "note"
sw.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
sw.row_dimensions[r].height = 58
sw.cell(row=r + 2, column=1, value=DISCLOSURE).style = "note"

# ===================================================== 6. Capital Structures ==
cw = wb.create_sheet("Capital Structures")
cw.column_dimensions["A"].width = 46
for i in range(2, 5):
    cw.column_dimensions[get_column_letter(i)].width = 20
r = sheet_header(cw, "Capital Structures",
                 "The same capital requirement under three structures. The operating forecast is identical in every case.",
                 "D")

cw.cell(row=r, column=1, value="Term").style = "hdr"
for i, name in enumerate(["Growth equity", "Private credit", "Blended capital"]):
    cw.cell(row=r, column=2 + i, value=name).style = "hdr"
r += 1

C = {}


def crow(label, key, fmt, f_eq, f_pc, f_bl, band=False):
    global r
    cw.cell(row=r, column=1, value=label).style = "lbl"
    C[key] = r
    for i, f in enumerate([f_eq, f_pc, f_bl]):
        c = cw.cell(row=r, column=2 + i, value=f)
        c.style = "tot" if band else "calc"
        c.number_format = fmt
    r += 1


crow("Total capital raised", "raise", FMT_M2, f"={A['raise']}", f"={A['raise']}", f"={A['raise']}")
crow("Equity component", "eq", FMT_M2, f"={A['eq_equity']}", "=0", f"={A['bl_equity']}")
crow("Debt component", "debt", FMT_M2, "=0", f"={A['pc_principal']}", f"={A['bl_debt']}")
crow("Original issue discount rate", "oidr", FMT_PCT, "=0", f"={A['pc_oid']}", f"={A['bl_oid']}")
crow("Original issue discount cost", "oid", FMT_M2, "=0",
     f"=C{C['debt']}*C{C['oidr']}", f"=D{C['debt']}*D{C['oidr']}")
crow("Net proceeds at close", "net", FMT_M2,
     f"=B{C['eq']}+B{C['debt']}-B{C['oid']}",
     f"=C{C['eq']}+C{C['debt']}-C{C['oid']}",
     f"=D{C['eq']}+D{C['debt']}-D{C['oid']}", band=True)
crow("Less existing debt refinanced", "refi", FMT_M2,
     f"=-{A['refi']}", f"=-{A['refi']}", f"=-{A['refi']}")
crow("Cash at close", "cash", FMT_M2,
     f"={A['cash0']}+B{C['net']}+B{C['refi']}",
     f"={A['cash0']}+C{C['net']}+C{C['refi']}",
     f"={A['cash0']}+D{C['net']}+D{C['refi']}", band=True)

r += 1
cw.cell(row=r, column=1, value="Ownership").style = "sec"
r += 1
crow("Pre-money valuation", "pre", FMT_M2,
     f"={A['beg_arr']}*{A['entryx']}", f"={A['beg_arr']}*{A['entryx']}", f"={A['beg_arr']}*{A['entryx']}")
crow("Post-money valuation", "post", FMT_M2,
     f"=B{C['pre']}+B{C['eq']}", f"=C{C['pre']}", f"=D{C['pre']}+D{C['eq']}")
crow("Investor ownership", "own", FMT_PCT,
     f"=B{C['eq']}/B{C['post']}", "=0", f"=D{C['eq']}/D{C['post']}")
crow("Founder and existing holder dilution", "dil", FMT_PCT,
     f"=B{C['own']}", f"=C{C['own']}", f"=D{C['own']}", band=True)

r += 1
cw.cell(row=r, column=1, value="Debt terms").style = "sec"
r += 1
crow("Cash interest rate", "ci", FMT_PCT, "=0", f"={A['pc_cash']}", f"={A['bl_cash']}")
crow("PIK interest rate", "pik", FMT_PCT, "=0", f"={A['pc_pik']}", f"={A['bl_pik']}")
crow("Maturity (years)", "mat", FMT_INT, "=0", f"={A['pc_mat']}", f"={A['bl_mat']}")
crow("Interest only period (years)", "io", FMT_INT, "=0", f"={A['pc_io']}", f"={A['bl_io']}")
crow("Annual amortisation (% of principal)", "am", FMT_PCT, "=0", f"={A['pc_amort']}", f"={A['bl_amort']}")

r += 1
cw.cell(row=r, column=1,
        value="Uses of proceeds are identical across all three structures: the existing term debt is refinanced in "
              "full at close and the balance funds product development, enterprise sales hiring, international "
              "expansion, working capital, and selective acquisitions. Holding uses constant is what makes the "
              "three structures comparable.").style = "note"
cw.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
cw.row_dimensions[r].height = 50
cw.cell(row=r + 2, column=1, value=DISCLOSURE).style = "note"

# ------------------------------------------------------------ case builder --
def build_case(sheet_name, title, subtitle, equity_ref, principal_ref, ci_ref,
               pik_ref, io_ref, amort_ref, oid_ref):
    """Create a case sheet with a debt schedule, cash roll forward, and returns."""
    w = wb.create_sheet(sheet_name)
    widths(w, 44, 15, 6)
    row = sheet_header(w, title, subtitle, "F")
    K = {}

    has_debt = principal_ref is not None

    # -- debt schedule ------------------------------------------------------
    if has_debt:
        w.cell(row=row, column=1, value="Debt schedule").style = "sec"
        row += 1
        row = year_headers(w, row, label="Line item (USD millions)")

        def drow(label, key, fmt, formula, band=False):
            nonlocal row
            w.cell(row=row, column=1, value=label).style = "lbl"
            K[key] = row
            for i in range(5):
                col = get_column_letter(2 + i)
                c = w.cell(row=row, column=2 + i, value=formula(i + 1, col))
                c.style = "tot" if band else "calc"
                c.number_format = fmt
            row += 1

        # Reserved, then filled once Closing balance exists.
        drow("Opening balance", "open", FMT_M2, lambda i, c: f"={principal_ref}")
        drow("Cash interest", "ci", FMT_M2, lambda i, c: f"={c}{K['open']}*{ci_ref}")
        drow("PIK interest", "pik", FMT_M2, lambda i, c: f"={c}{K['open']}*{pik_ref}")
        drow("Amortisation", "am", FMT_M2,
             lambda i, c: f"=IF({i}>{io_ref},{principal_ref}*{amort_ref},0)")
        drow("Debt service", "ds", FMT_M2, lambda i, c: f"={c}{K['ci']}+{c}{K['am']}", band=True)
        drow("Closing balance", "close", FMT_M2,
             lambda i, c: f"={c}{K['open']}+{c}{K['pik']}-{c}{K['am']}", band=True)
        # Rewrite opening balance now that the closing row exists.
        for i in range(5):
            prev = get_column_letter(1 + i)
            c = w.cell(row=K["open"], column=2 + i,
                       value=f"={principal_ref}" if i == 0 else f"={prev}{K['close']}")
            c.style = "calc"
            c.number_format = FMT_M2

        drow("Debt to ending ARR", "lev", FMT_X,
             lambda i, c: f"={c}{K['close']}/'Operating Forecast'!{c}{O['end']}")
        drow("Interest coverage (EBITDA / cash interest)", "icov", FMT_X,
             lambda i, c: f"=IF({c}{K['ci']}=0,\"n/a\",'Operating Forecast'!{c}{O['ebitda']}/{c}{K['ci']})")
        drow("Debt service coverage ratio", "dscr", FMT_X,
             lambda i, c: f"=IF({c}{K['ds']}=0,\"n/a\",('Operating Forecast'!{c}{O['ebitda']}+'Operating Forecast'!{c}{O['capex']})/{c}{K['ds']})")

        w.conditional_formatting.add(
            f"B{K['icov']}:F{K['dscr']}",
            CellIsRule(operator="lessThan", formula=["1"], font=Font(color="9B1C1C")),
        )
        row += 1

    # -- cash roll forward ---------------------------------------------------
    w.cell(row=row, column=1, value="Cash roll forward").style = "sec"
    row += 1
    row = year_headers(w, row, label="Line item (USD millions)")

    def krow(label, key, fmt, formula, band=False):
        nonlocal row
        w.cell(row=row, column=1, value=label).style = "lbl"
        K[key] = row
        for i in range(5):
            col = get_column_letter(2 + i)
            c = w.cell(row=row, column=2 + i, value=formula(i + 1, col))
            c.style = "tot" if band else "calc"
            c.number_format = fmt
        row += 1

    cash_col = {"Growth Equity Case": "B", "Private Credit Case": "C", "Blended Capital Case": "D"}[sheet_name]
    # Reserved, then filled once Closing cash exists.
    krow("Opening cash", "copen", FMT_M2,
         lambda i, c: f"='Capital Structures'!{cash_col}{C['cash']}")
    krow("Unlevered free cash flow", "cfcf", FMT_M2,
         lambda i, c: f"='Operating Forecast'!{c}{O['fcf']}")
    krow("Cash interest", "cci", FMT_M2,
         lambda i, c: f"=-{c}{K['ci']}" if has_debt else "=0")
    krow("Amortisation", "cam", FMT_M2,
         lambda i, c: f"=-{c}{K['am']}" if has_debt else "=0")
    krow("Closing cash", "cclose", FMT_M2,
         lambda i, c: f"=SUM({c}{K['copen']}:{c}{K['cam']})", band=True)
    for i in range(5):
        prev = get_column_letter(1 + i)
        c = w.cell(row=K["copen"], column=2 + i,
                   value=f"='Capital Structures'!{cash_col}{C['cash']}" if i == 0
                   else f"={prev}{K['cclose']}")
        c.style = "calc"
        c.number_format = FMT_M2

    krow("Minimum cash covenant", "mincash", FMT_M2, lambda i, c: f"={A['mincash']}")
    krow("Covenant headroom", "head", FMT_M2,
         lambda i, c: f"={c}{K['cclose']}-{c}{K['mincash']}")
    krow("Covenant status", "status", "General",
         lambda i, c: f'=IF({c}{K["cclose"]}<{c}{K["mincash"]},"BREACH","Headroom")')
    krow("Runway (months at current burn)", "runway", '0.0',
         lambda i, c: f'=IF({c}{K["cclose"]}-{c}{K["copen"]}>=0,"cash generative",{c}{K["cclose"]}/({c}{K["copen"]}-{c}{K["cclose"]})*12)')

    w.conditional_formatting.add(
        f"B{K['cclose']}:F{K['cclose']}",
        CellIsRule(operator="lessThan", formula=[f"{A['mincash']}"],
                   fill=PatternFill("solid", bgColor="FDE2E2"), font=Font(color="9B1C1C", bold=True)),
    )
    w.conditional_formatting.add(
        f"B{K['status']}:F{K['status']}",
        CellIsRule(operator="equal", formula=['"BREACH"'],
                   fill=PatternFill("solid", bgColor="FDE2E2"), font=Font(color="9B1C1C", bold=True)),
    )
    row += 1

    # -- returns -------------------------------------------------------------
    w.cell(row=row, column=1, value="Exit and returns").style = "sec"
    row += 1

    def vrow(label, key, fmt, formula, comment=None, band=False):
        nonlocal row
        w.cell(row=row, column=1, value=label).style = "lbl"
        K[key] = row
        c = w.cell(row=row, column=2, value=formula)
        c.style = "tot" if band else "calc"
        c.number_format = fmt
        if comment:
            c.comment = Comment(comment, "Model guide")
        row += 1

    vrow("Year 5 ending ARR", "exarr", FMT_M2, f"='Operating Forecast'!F{O['end']}")
    vrow("Exit ARR multiple", "exx", FMT_X, f"={A['exitx']}")
    vrow("Exit enterprise value", "exev", FMT_M2, f"=B{K['exarr']}*B{K['exx']}")
    vrow("Plus year 5 cash", "excash", FMT_M2, f"=F{K['cclose']}")
    vrow("Less year 5 debt", "exdebt", FMT_M2, f"=-F{K['close']}" if has_debt else "=0")
    vrow("Exit equity value", "exeq", FMT_M2,
         f"=B{K['exev']}+B{K['excash']}+B{K['exdebt']}", band=True)

    own_col = {"Growth Equity Case": "B", "Private Credit Case": "C", "Blended Capital Case": "D"}[sheet_name]
    vrow("Equity investor ownership", "own", FMT_PCT, f"='Capital Structures'!{own_col}{C['own']}")
    vrow("Equity invested", "eqin", FMT_M2, f"={equity_ref}" if equity_ref else "=0")
    vrow("Equity investor proceeds at exit", "eqout", FMT_M2, f"=B{K['exeq']}*B{K['own']}")
    vrow("Equity investor MOIC", "moic", FMT_X,
         f'=IF(B{K["eqin"]}=0,"n/a",B{K["eqout"]}/B{K["eqin"]})', band=True)
    vrow("Equity investor IRR", "irr", FMT_PCT,
         f'=IF(B{K["eqin"]}=0,"n/a",IRR({{0}}))'.replace("{0}", f"B{row+2}:G{row+2}") if False else
         f'=IF(B{K["eqin"]}=0,"n/a",(B{K["eqout"]}/B{K["eqin"]})^(1/5)-1)',
         "Five year single entry and single exit, so the IRR reduces to the fifth root of MOIC less one. "
         "An explicit cash flow line is laid out on the Returns Analysis sheet.", band=True)

    if has_debt:
        row += 1
        w.cell(row=row, column=1, value="Debt investor return").style = "sec"
        row += 1
        vrow("Principal advanced", "dprin", FMT_M2, f"={principal_ref}")
        vrow("Less original issue discount", "doid", FMT_M2, f"=-{principal_ref}*{oid_ref}")
        vrow("Net outlay at close", "dout", FMT_M2, f"=B{K['dprin']}+B{K['doid']}", band=True)
        vrow("Total cash interest received", "dci", FMT_M2, f"=SUM(B{K['ci']}:F{K['ci']})")
        vrow("Total amortisation received", "dam", FMT_M2, f"=SUM(B{K['am']}:F{K['am']})")
        vrow("Balance repaid at maturity", "dbal", FMT_M2, f"=F{K['close']}")
        vrow("Total received", "dtot", FMT_M2,
             f"=B{K['dci']}+B{K['dam']}+B{K['dbal']}", band=True)
        vrow("Debt investor MOIC", "dmoic", FMT_X, f"=B{K['dtot']}/B{K['dout']}", band=True)
        vrow("Refinancing risk at maturity", "drefi", "General",
             f'=IF(F{K["close"]}>0,"Yes, balance of "&TEXT(F{K["close"]},"0.0")&"m outstanding at maturity","No")')

    w.cell(row=row + 1, column=1, value=DISCLOSURE).style = "note"
    return w, K


eq_sheet, EQ = build_case(
    "Growth Equity Case",
    "Growth Equity Case",
    f"USD 20 million of primary equity. No leverage. The safest structure and the most expensive in ownership terms.",
    A["eq_equity"], None, None, None, None, None, None,
)

pc_sheet, PC = build_case(
    "Private Credit Case",
    "Private Credit Case",
    "USD 20 million senior secured. No dilution, and the company cannot service it from operating cash flow.",
    None, A["pc_principal"], A["pc_cash"], A["pc_pik"], A["pc_io"], A["pc_amort"], A["pc_oid"],
)

bl_sheet, BL = build_case(
    "Blended Capital Case",
    "Blended Capital Case",
    "USD 8 million of primary equity alongside a USD 12 million senior secured facility. The recommended structure.",
    A["bl_equity"], A["bl_debt"], A["bl_cash"], A["bl_pik"], A["bl_io"], A["bl_amort"], A["bl_oid"],
)

# ========================================================= 10. Debt Schedule ==
dw = wb.create_sheet("Debt Schedule")
widths(dw, 46, 15, 6)
r = sheet_header(dw, "Debt Schedule",
                 "Both debt structures side by side, drawn from the case sheets. Terms are editable on the Assumptions sheet.",
                 "F")

for label, sheet, K in (("Private credit: USD 20 million facility", "Private Credit Case", PC),
                        ("Blended capital: USD 12 million facility", "Blended Capital Case", BL)):
    dw.cell(row=r, column=1, value=label).style = "sec"
    dw.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    r = year_headers(dw, r, label="Line item (USD millions)")
    for name, key, fmt in (
        ("Opening balance", "open", FMT_M2),
        ("Cash interest", "ci", FMT_M2),
        ("PIK interest", "pik", FMT_M2),
        ("Amortisation", "am", FMT_M2),
        ("Debt service", "ds", FMT_M2),
        ("Closing balance", "close", FMT_M2),
        ("Debt to ending ARR", "lev", FMT_X),
        ("Interest coverage", "icov", FMT_X),
        ("Debt service coverage ratio", "dscr", FMT_X),
    ):
        dw.cell(row=r, column=1, value=name).style = "lbl"
        for i in range(5):
            col = get_column_letter(2 + i)
            c = dw.cell(row=r, column=2 + i, value=f"='{sheet}'!{col}{K[key]}")
            c.style = "tot" if key in ("ds", "close") else "calc"
            c.number_format = fmt
        r += 1
    r += 1

dw.cell(row=r, column=1,
        value="Interest coverage is EBITDA divided by cash interest and is negative for as long as EBITDA is "
              "negative. Debt service coverage is EBITDA less capital expenditure over total debt service. "
              "Neither ratio reaches 1.0x on the USD 20 million facility inside the forecast period, which is "
              "the arithmetic reason the all debt structure is rejected.").style = "note"
dw.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
dw.row_dimensions[r].height = 50
dw.cell(row=r + 2, column=1, value=DISCLOSURE).style = "note"

# ====================================================== 11. Returns Analysis ==
rw = wb.create_sheet("Returns Analysis")
rw.column_dimensions["A"].width = 46
for i in range(2, 9):
    rw.column_dimensions[get_column_letter(i)].width = 15
r = sheet_header(rw, "Returns Analysis",
                 "Explicit cash flows and returns for equity and debt investors under each structure.", "G")

rw.cell(row=r, column=1, value="Equity investor cash flows (USD millions)").style = "sec"
rw.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
r += 1
rw.cell(row=r, column=1, value="Structure").style = "hdr"
for i, y in enumerate(["Close", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]):
    rw.cell(row=r, column=2 + i, value=y).style = "hdr"
r += 1

eq_flow_rows = {}
for name, sheet, K in (("Growth equity", "Growth Equity Case", EQ),
                       ("Blended capital", "Blended Capital Case", BL)):
    rw.cell(row=r, column=1, value=name).style = "lbl"
    c = rw.cell(row=r, column=2, value=f"=-'{sheet}'!B{K['eqin']}")
    c.style = "calc"
    c.number_format = FMT_M2
    for i in range(1, 5):
        c = rw.cell(row=r, column=2 + i, value=0)
        c.style = "calc"
        c.number_format = FMT_M2
    c = rw.cell(row=r, column=7, value=f"='{sheet}'!B{K['eqout']}")
    c.style = "tot"
    c.number_format = FMT_M2
    eq_flow_rows[name] = r
    r += 1

r += 1
rw.cell(row=r, column=1, value="Equity investor returns").style = "sec"
rw.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
r += 1
rw.cell(row=r, column=1, value="Metric").style = "hdr"
for i, name in enumerate(["Growth equity", "Private credit", "Blended capital"]):
    rw.cell(row=r, column=2 + i, value=name).style = "hdr"
r += 1

R = {}


def rrow(label, key, fmt, vals, band=False):
    global r
    rw.cell(row=r, column=1, value=label).style = "lbl"
    R[key] = r
    for i, v in enumerate(vals):
        c = rw.cell(row=r, column=2 + i, value=v)
        c.style = "tot" if band else "calc"
        c.number_format = fmt
    r += 1


rrow("Equity invested", "eqin", FMT_M2,
     [f"='Growth Equity Case'!B{EQ['eqin']}", "=0", f"='Blended Capital Case'!B{BL['eqin']}"])
rrow("Investor ownership", "own", FMT_PCT,
     [f"='Growth Equity Case'!B{EQ['own']}", "=0", f"='Blended Capital Case'!B{BL['own']}"])
rrow("Exit equity value", "exeq", FMT_M2,
     [f"='Growth Equity Case'!B{EQ['exeq']}", f"='Private Credit Case'!B{PC['exeq']}",
      f"='Blended Capital Case'!B{BL['exeq']}"])
rrow("Equity investor proceeds", "eqout", FMT_M2,
     [f"='Growth Equity Case'!B{EQ['eqout']}", "=0", f"='Blended Capital Case'!B{BL['eqout']}"])
rrow("MOIC", "moic", FMT_X,
     [f"=IRR" if False else f"='Growth Equity Case'!B{EQ['moic']}", '="n/a"',
      f"='Blended Capital Case'!B{BL['moic']}"], band=True)
rrow("IRR (5 year hold)", "irr", FMT_PCT,
     [f"=IRR(B{eq_flow_rows['Growth equity']}:G{eq_flow_rows['Growth equity']})", '="n/a"',
      f"=IRR(B{eq_flow_rows['Blended capital']}:G{eq_flow_rows['Blended capital']})"], band=True)
rrow("Founder and existing holder dilution", "dil", FMT_PCT,
     [f"='Capital Structures'!B{C['dil']}", f"='Capital Structures'!C{C['dil']}",
      f"='Capital Structures'!D{C['dil']}"])

r += 1
rw.cell(row=r, column=1, value="Debt investor returns").style = "sec"
rw.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
r += 1
rw.cell(row=r, column=1, value="Metric").style = "hdr"
for i, name in enumerate(["Growth equity", "Private credit", "Blended capital"]):
    rw.cell(row=r, column=2 + i, value=name).style = "hdr"
r += 1

rrow("Net outlay at close", "dout", FMT_M2,
     ['="n/a"', f"='Private Credit Case'!B{PC['dout']}", f"='Blended Capital Case'!B{BL['dout']}"])
rrow("Total received over five years", "dtot", FMT_M2,
     ['="n/a"', f"='Private Credit Case'!B{PC['dtot']}", f"='Blended Capital Case'!B{BL['dtot']}"])
rrow("Debt investor MOIC", "dmoic", FMT_X,
     ['="n/a"', f"='Private Credit Case'!B{PC['dmoic']}", f"='Blended Capital Case'!B{BL['dmoic']}"], band=True)

# Explicit debt cash flow rows so IRR is computed from flows rather than asserted.
r += 1
rw.cell(row=r, column=1, value="Debt investor cash flows (USD millions)").style = "sec"
rw.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
r += 1
rw.cell(row=r, column=1, value="Structure").style = "hdr"
for i, y in enumerate(["Close", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]):
    rw.cell(row=r, column=2 + i, value=y).style = "hdr"
r += 1

debt_flow_rows = {}
for name, sheet, K in (("Private credit", "Private Credit Case", PC),
                       ("Blended capital", "Blended Capital Case", BL)):
    rw.cell(row=r, column=1, value=name).style = "lbl"
    c = rw.cell(row=r, column=2, value=f"=-'{sheet}'!B{K['dout']}")
    c.style = "calc"
    c.number_format = FMT_M2
    for i in range(5):
        col = get_column_letter(2 + i)
        formula = (f"='{sheet}'!{col}{K['ci']}+'{sheet}'!{col}{K['am']}"
                   + (f"+'{sheet}'!{col}{K['close']}" if i == 4 else ""))
        c = rw.cell(row=r, column=3 + i, value=formula)
        c.style = "tot" if i == 4 else "calc"
        c.number_format = FMT_M2
    debt_flow_rows[name] = r
    r += 1

# Columns align with the Growth equity / Private credit / Blended capital header
# above, so the growth equity column is deliberately left as not applicable.
rw.cell(row=r, column=1, value="Debt investor IRR").style = "lbl"
c = rw.cell(row=r, column=2, value='="n/a"')
c.style = "tot"
for i, name in enumerate(["Private credit", "Blended capital"]):
    c = rw.cell(row=r, column=3 + i,
                value=f"=IRR(B{debt_flow_rows[name]}:G{debt_flow_rows[name]})")
    c.style = "tot"
    c.number_format = FMT_PCT
r += 1

r += 1
rw.cell(row=r, column=1,
        value="The equity investor MOIC is higher under the blended structure than under all equity, because a "
              "smaller cheque is exposed to the same enterprise value appreciation. That is the return argument "
              "for the blend. The dilution argument is stronger still, and the covenant argument is what makes "
              "the all debt structure unavailable.").style = "note"
rw.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
rw.row_dimensions[r].height = 50
rw.cell(row=r + 2, column=1, value=DISCLOSURE).style = "note"

# ========================================================= 12. Sensitivities ==
zw = wb.create_sheet("Sensitivities")
zw.column_dimensions["A"].width = 34
for i in range(2, 9):
    zw.column_dimensions[get_column_letter(i)].width = 15
r = sheet_header(zw, "Sensitivities",
                 "Each table moves one driver and holds everything else constant, so the effect is attributable.",
                 "G")


def sens_header(row, title, note=None):
    zw.cell(row=row, column=1, value=title).style = "sec"
    zw.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1
    if note:
        c = zw.cell(row=row, column=1, value=note)
        c.style = "note"
        zw.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        zw.row_dimensions[row].height = 26
        row += 1
    return row


# --- Exit multiple ---------------------------------------------------------
r = sens_header(r, "Exit multiple sensitivity, blended capital equity investor",
                "Year five ARR and ending cash are held at the base case. Only the exit ARR multiple moves.")
zw.cell(row=r, column=1, value="Exit ARR multiple").style = "hdr"
for i, h in enumerate(["Enterprise value", "Exit equity value", "Proceeds", "MOIC", "IRR"]):
    zw.cell(row=r, column=2 + i, value=h).style = "hdr"
r += 1
exit_start = r
for i, mult in enumerate([5.0, 6.0, 7.0, 8.0, 9.0]):
    c = zw.cell(row=r, column=1, value=mult)
    c.style = "inp"
    c.number_format = FMT_X
    ev = zw.cell(row=r, column=2, value=f"='Blended Capital Case'!B{BL['exarr']}*A{r}")
    ev.style = "calc"; ev.number_format = FMT_M2
    eq = zw.cell(row=r, column=3,
                 value=f"=B{r}+'Blended Capital Case'!B{BL['excash']}+'Blended Capital Case'!B{BL['exdebt']}")
    eq.style = "calc"; eq.number_format = FMT_M2
    pr = zw.cell(row=r, column=4, value=f"=C{r}*'Blended Capital Case'!B{BL['own']}")
    pr.style = "calc"; pr.number_format = FMT_M2
    mo = zw.cell(row=r, column=5, value=f"=D{r}/'Blended Capital Case'!B{BL['eqin']}")
    mo.style = "tot"; mo.number_format = FMT_X
    ir = zw.cell(row=r, column=6, value=f"=E{r}^(1/5)-1")
    ir.style = "tot"; ir.number_format = FMT_PCT
    r += 1
r += 1

# --- ARR growth ------------------------------------------------------------
r = sens_header(r, "ARR growth sensitivity",
                "Year one growth varies. The same annual decay applies in later years.")
zw.cell(row=r, column=1, value="Year 1 ARR growth").style = "hdr"
for i, h in enumerate(["Year 5 ending ARR", "Exit enterprise value at base multiple"]):
    zw.cell(row=r, column=2 + i, value=h).style = "hdr"
r += 1
for g in [0.20, 0.25, 0.30, 0.35, 0.40]:
    c = zw.cell(row=r, column=1, value=g)
    c.style = "inp"; c.number_format = FMT_PCT
    arr = (f"={A['beg_arr']}"
           + "".join(f"*(1+MAX(0,A{r}-{A['decay']}*{k}))" for k in range(5)))
    a = zw.cell(row=r, column=2, value=arr)
    a.style = "calc"; a.number_format = FMT_M2
    e = zw.cell(row=r, column=3, value=f"=B{r}*{A['exitx']}")
    e.style = "tot"; e.number_format = FMT_M2
    r += 1
r += 1

# --- Retention -------------------------------------------------------------
r = sens_header(r, "Retention sensitivity",
                "Net revenue retention varies while the ARR growth target is held fixed, so lower retention "
                "simply raises the New ARR the sales team must win.")
zw.cell(row=r, column=1, value="Net revenue retention").style = "hdr"
for i, h in enumerate(["Year 1 New ARR required", "Increase over base case", "Implied year 1 CAC payback (months)"]):
    zw.cell(row=r, column=2 + i, value=h).style = "hdr"
r += 1
ret_start = r
for n in [1.00, 1.05, 1.10, 1.15, 1.20]:
    c = zw.cell(row=r, column=1, value=n)
    c.style = "inp"; c.number_format = FMT_PCT
    a = zw.cell(row=r, column=2,
                value=f"={A['beg_arr']}*(1+'Operating Forecast'!B{O['g']})-{A['beg_arr']}*A{r}")
    a.style = "calc"; a.number_format = FMT_M2
    d = zw.cell(row=r, column=3, value=f"=B{r}-'Operating Forecast'!B{O['new']}")
    d.style = "calc"; d.number_format = FMT_M2
    p = zw.cell(row=r, column=4,
                value=f"=-'Operating Forecast'!B{O['sm']}/((B{r}+{A['beg_arr']}*(A{r}-{A['grr']}))*{A['gm']})*12")
    p.style = "tot"; p.number_format = '0.0'
    r += 1
r += 1

# --- Interest rate ---------------------------------------------------------
r = sens_header(r, "Interest rate sensitivity, blended capital facility",
                "Applied to the blended facility principal. Shows where the minimum cash covenant starts to bind.")
zw.cell(row=r, column=1, value="Cash interest rate").style = "hdr"
for i, h in enumerate(["Year 1 cash interest", "Five year cash interest", "Year 5 closing cash", "Covenant status"]):
    zw.cell(row=r, column=2 + i, value=h).style = "hdr"
r += 1
ir_start = r
for rate in [0.09, 0.10, 0.11, 0.12, 0.13]:
    c = zw.cell(row=r, column=1, value=rate)
    c.style = "inp"; c.number_format = FMT_PCT
    y1 = zw.cell(row=r, column=2, value=f"={A['bl_debt']}*A{r}")
    y1.style = "calc"; y1.number_format = FMT_M2
    # Five years of interest on a balance that amortises after the interest only period.
    tot = zw.cell(row=r, column=3,
                  value=f"={A['bl_debt']}*A{r}*{A['bl_io']}"
                        f"+({A['bl_debt']}*A{r}+({A['bl_debt']}-{A['bl_debt']}*{A['bl_amort']})*A{r})")
    tot.style = "calc"; tot.number_format = FMT_M2
    cash = zw.cell(row=r, column=4,
                   value=f"='Blended Capital Case'!F{BL['cclose']}"
                         f"+(SUM('Blended Capital Case'!B{BL['ci']}:F{BL['ci']})-C{r})")
    cash.style = "tot"; cash.number_format = FMT_M2
    st = zw.cell(row=r, column=5,
                 value=f'=IF(D{r}<{A["mincash"]},"BREACH","Headroom")')
    st.style = "calc"
    r += 1

zw.conditional_formatting.add(
    f"E{ir_start}:E{r-1}",
    CellIsRule(operator="equal", formula=['"BREACH"'],
               fill=PatternFill("solid", bgColor="FDE2E2"), font=Font(color="9B1C1C", bold=True)),
)
r += 1

# --- Break even ------------------------------------------------------------
r = sens_header(r, "Break even analysis",
                "At what point the cost structure stops consuming cash.")
be = [
    ("Year 5 operating expenses as a percentage of revenue", FMT_PCT,
     f"=-'Operating Forecast'!F{O['opex']}/'Operating Forecast'!F{O['rev']}"),
    ("Gross margin", FMT_PCT, f"={A['gm']}"),
    ("Contribution margin at the year 5 cost structure", FMT_PCT,
     f"={A['gm']}+'Operating Forecast'!F{O['opex']}/'Operating Forecast'!F{O['rev']}"),
    ("First year of positive EBITDA", "General",
     f"=IF('Operating Forecast'!B{O['ebitda']}>0,1,IF('Operating Forecast'!C{O['ebitda']}>0,2,"
     f"IF('Operating Forecast'!D{O['ebitda']}>0,3,IF('Operating Forecast'!E{O['ebitda']}>0,4,"
     f"IF('Operating Forecast'!F{O['ebitda']}>0,5,\"beyond year 5\")))))"),
    ("First year of positive unlevered free cash flow", "General",
     f"=IF('Operating Forecast'!B{O['fcf']}>0,1,IF('Operating Forecast'!C{O['fcf']}>0,2,"
     f"IF('Operating Forecast'!D{O['fcf']}>0,3,IF('Operating Forecast'!E{O['fcf']}>0,4,"
     f"IF('Operating Forecast'!F{O['fcf']}>0,5,\"beyond year 5\")))))"),
    ("Cumulative unlevered free cash flow over five years", FMT_M2,
     f"=SUM('Operating Forecast'!B{O['fcf']}:F{O['fcf']})"),
    ("Cumulative cash burn before financing", FMT_M2,
     f"=-SUM('Operating Forecast'!B{O['fcf']}:F{O['fcf']})"),
]
for label, fmt, formula in be:
    zw.cell(row=r, column=1, value=label).style = "lbl"
    c = zw.cell(row=r, column=2, value=formula)
    c.style = "calc"
    if fmt != "General":
        c.number_format = fmt
    r += 1

zw.cell(row=r + 1, column=1, value=DISCLOSURE).style = "note"

# ========================================================= 13. Downside Case ==
xw = wb.create_sheet("Downside Case")
widths(xw, 46, 15, 6)
r = sheet_header(xw, "Downside Case",
                 "Lower growth, no net expansion, higher churn, thinner margin, and sales spend that does not fall. "
                 "This is the test that separates the three structures.",
                 "F")
r = year_headers(xw, r, label="Line item (USD millions)")

D = {}


def drow2(name, key, fmt, formula, band=False):
    global r
    xw.cell(row=r, column=1, value=name).style = "lbl"
    D[key] = r
    for i in range(5):
        col = get_column_letter(2 + i)
        c = xw.cell(row=r, column=2 + i, value=formula(i + 1, col))
        c.style = "tot" if band else "calc"
        c.number_format = fmt
    r += 1


# Reserved, then filled once Ending ARR exists.
drow2("Beginning ARR", "beg", FMT_M2, lambda i, c: f"={A['beg_arr']}")
drow2("Growth rate", "g", FMT_PCT, lambda i, c: f"=MAX(0,{A['ds_g1']}-{A['decay']}*{i-1})")
drow2("Ending ARR", "end", FMT_M2, lambda i, c: f"={c}{D['beg']}*(1+{c}{D['g']})", band=True)
for i in range(5):
    prev = get_column_letter(1 + i)
    c = xw.cell(row=D["beg"], column=2 + i,
                value=f"={A['beg_arr']}" if i == 0 else f"={prev}{D['end']}")
    c.style = "calc"; c.number_format = FMT_M2

drow2("New ARR required", "new", FMT_M2,
      lambda i, c: f"={c}{D['end']}-{c}{D['beg']}*{A['ds_nrr']}")
drow2("Churned ARR", "chn", FMT_M2, lambda i, c: f"=-{c}{D['beg']}*(1-{A['ds_grr']})")
drow2("Revenue", "rev", FMT_M2, lambda i, c: f"=({c}{D['beg']}+{c}{D['end']})/2")
drow2("Gross profit", "gp", FMT_M2, lambda i, c: f"={c}{D['rev']}*{A['ds_gm']}")
drow2("Total operating expenses", "opex", FMT_M2,
      lambda i, c: f"=-{c}{D['rev']}*(Assumptions!{c}${A['sm_row']}+{A['ds_smup']}"
                   f"+Assumptions!{c}${A['rd_row']}+Assumptions!{c}${A['ga_row']})")
drow2("EBITDA", "ebitda", FMT_M2, lambda i, c: f"={c}{D['gp']}+{c}{D['opex']}", band=True)
drow2("Capital expenditure", "capex", FMT_M2, lambda i, c: f"=-{c}{D['rev']}*{A['capex']}")
drow2("Deferred revenue benefit", "drb", FMT_M2,
      lambda i, c: f"=({c}{D['end']}-{c}{D['beg']})*{A['dr']}")
drow2("Unlevered free cash flow", "fcf", FMT_M2,
      lambda i, c: f"={c}{D['ebitda']}+{c}{D['capex']}+{c}{D['drb']}", band=True)

xw.conditional_formatting.add(
    f"B{D['ebitda']}:F{D['ebitda']}",
    CellIsRule(operator="lessThan", formula=["0"], font=Font(color="9B1C1C", bold=True)),
)

r += 1
xw.cell(row=r, column=1, value="Cash outcome under each structure").style = "sec"
xw.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
r += 1
r = year_headers(xw, r, label="Closing cash (USD millions)")

ds_rows = {}
for label, cash_col, K in (("Growth equity", "B", EQ),
                           ("Private credit", "C", PC),
                           ("Blended capital", "D", BL)):
    xw.cell(row=r, column=1, value=label).style = "lbl"
    has_debt = "ci" in K and label != "Growth equity"
    for i in range(5):
        col = get_column_letter(2 + i)
        prev = get_column_letter(1 + i)
        base = (f"='Capital Structures'!{cash_col}{C['cash']}" if i == 0 else f"={prev}{r}")
        svc = ""
        if has_debt:
            sheet = "Private Credit Case" if label == "Private credit" else "Blended Capital Case"
            svc = f"-'{sheet}'!{col}{K['ci']}-'{sheet}'!{col}{K['am']}"
        c = xw.cell(row=r, column=2 + i, value=f"{base}+{col}{D['fcf']}{svc}")
        c.style = "calc"
        c.number_format = FMT_M2
    ds_rows[label] = r
    r += 1

for label, row_i in ds_rows.items():
    xw.conditional_formatting.add(
        f"B{row_i}:F{row_i}",
        CellIsRule(operator="lessThan", formula=[f"{A['mincash']}"],
                   fill=PatternFill("solid", bgColor="FDE2E2"), font=Font(color="9B1C1C", bold=True)),
    )

r += 1
xw.cell(row=r, column=1, value="Structure").style = "hdr"
for i, h in enumerate(["Year 5 cash", "Survives five years", "Assessment"]):
    xw.cell(row=r, column=2 + i, value=h).style = "hdr"
r += 1
for label, row_i in ds_rows.items():
    xw.cell(row=r, column=1, value=label).style = "lbl"
    c = xw.cell(row=r, column=2, value=f"=F{row_i}")
    c.style = "tot"; c.number_format = FMT_M2
    c = xw.cell(row=r, column=3, value=f'=IF(F{row_i}>0,"Yes","No")')
    c.style = "calc"
    c = xw.cell(row=r, column=4,
                value=f'=IF(F{row_i}<=0,"Fails: runs out of cash",'
                      f'IF(MIN(B{row_i}:F{row_i})<{A["mincash"]},"Survives but breaches the minimum cash covenant",'
                      f'"Survives with covenant headroom"))')
    c.style = "calc"
    r += 1

r += 1
xw.cell(row=r, column=1,
        value="This is the honest boundary of the recommendation. In the base case the blended structure is "
              "clearly best. In the downside case it survives only narrowly and breaches the minimum cash "
              "covenant, while the all debt structure fails outright and the all equity structure is comfortable. "
              "A reader who assigns a high probability to the downside should prefer more equity in the mix, and "
              "the equity component is an input on the Assumptions sheet precisely so that view can be tested.").style = "note"
xw.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
xw.row_dimensions[r].height = 62
xw.cell(row=r + 2, column=1, value=DISCLOSURE).style = "note"

# ============================================== 14. Sources and Disclosures ==
vw = wb.create_sheet("Sources and Disclosures")
vw.column_dimensions["A"].width = 30
vw.column_dimensions["B"].width = 108
vw["A1"] = "Sources and Disclosures"
vw["A1"].style = "title"
vw["A2"] = "What this workbook is, what it is not, and where its inputs come from."
vw["A2"].style = "sub"

r = 4
disclosures = [
    ("Hypothetical company", LONG_DISCLOSURE),
    ("Illustrative figures",
     "All operating, financial, and transaction assumptions in this workbook are illustrative. They do not "
     "represent an actual company, an actual financing, or an investment recommendation. Every output sheet "
     "carries the line: All figures are hypothetical and illustrative."),
    ("Project disclosure",
     "This is an independent work sample built by Sahil Modi. It is not affiliated with or endorsed by any "
     "investment firm. The private company universe published alongside this model is based on dated public "
     "sources, and missing information there is identified as not publicly disclosed."),
    ("Source of the assumptions",
     "The assumptions were chosen to represent a plausible mid stage B2B enterprise SaaS company at the point "
     "where a growth financing decision becomes unavoidable: roughly USD 12 million of ARR, growth in the low "
     "thirties, net revenue retention above 100 percent, gross retention in the high eighties, gross margin "
     "near 80 percent, and negative EBITDA. They are not drawn from any real company."),
    ("Relationship to the company universe",
     "No company in the real private company universe appears in this workbook, and this hypothetical company "
     "appears nowhere in that universe. The separation is enforced by automated tests in the repository."),
    ("Debt terms",
     "Pricing, tenor, amortisation, original issue discount, and covenant levels are illustrative and set at "
     "levels broadly consistent with senior secured growth facilities for software companies. They are not "
     "quotes and do not reflect any lender's actual terms."),
    ("What has been deliberately excluded",
     "No estimate of any real company's ARR, revenue, retention, gross margin, EBITDA, burn, runway, valuation, "
     "customer concentration, debt balance, covenants, profitability, founder ownership, or exit value appears "
     "anywhere in this workbook or in the accompanying research."),
    ("Development disclosure",
     "Sahil Modi designed the research framework, scoring logic, underwriting structure, and investment "
     "analysis. AI-assisted development tools were used to support coding, research organization, testing, and "
     "document production. Every company record and material claim was reviewed against dated public sources."),
    ("Repository",
     "Source code, the company research records, and the script that generates this workbook are public at "
     "https://github.com/smodi13/growth-capital-origination-engine"),
    ("Contact", "Sahil Modi, modi.sahil@gmail.com"),
]
for label, text in disclosures:
    vw.cell(row=r, column=1, value=label).style = "sec"
    c = vw.cell(row=r, column=2, value=text)
    c.style = "body"
    vw.row_dimensions[r].height = max(30, 13 * (len(text) // 108 + 1))
    r += 1

# --------------------------------------------------------------- finalise --
for ws_ in wb.worksheets:
    ws_.sheet_view.showGridLines = False

wb.active = 0
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print(f"Wrote {OUT}")
print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
